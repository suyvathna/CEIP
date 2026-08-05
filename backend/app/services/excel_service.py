from io import BytesIO

from openpyxl import Workbook


def generate_project_report_excel(report: dict):
    workbook = Workbook()

    sheet = workbook.active
    sheet.title = "Project Report"

    sheet.append(["Item", "Value"])

    sheet.append(["Project Name", report["project_name"]])
    sheet.append(["Total Events", report["total_events"]])
    sheet.append(["Total Daily Logs", report["total_daily_logs"]])
    sheet.append(["Total Evidence", report["total_evidence"]])

    sheet.append(["Open Events", report["open_events"]])
    sheet.append(["Closed Events", report["closed_events"]])

    sheet.append(["High Severity", report["high_severity"]])
    sheet.append(["Medium Severity", report["medium_severity"]])
    sheet.append(["Low Severity", report["low_severity"]])

    sheet.append(["Latest Event", report["latest_event"]])

    sheet.append(["Generated At", str(report["generated_at"])])

    output = BytesIO()

    workbook.save(output)

    output.seek(0)

    return output


def _get(row, key):
    return row.get(key) if isinstance(row, dict) else getattr(row, key)


def _write_daily_log_sheet(sheet, report: dict):
    """
    One Daily Log's worth of rows, in the same section order the PDF
    export uses (see pdf_service._daily_log_flowables) - kept as its own
    function so both the single-day and project-wide compiled Excel
    exports build sheets the same way.
    """
    sheet.append([f"Daily Log: {report['diary_date']}"])
    sheet.append([])

    sheet.append(["WEATHER REPORT"])
    sheet.append(["Temp Avg", "Humidity Avg"])
    sheet.append(
        [
            report.get("temp_avg_c"),
            report.get("humidity_avg_pct"),
        ]
    )
    sheet.append([])

    snapshot = report.get("daily_snapshot") or []
    if snapshot:
        sheet.append(["DAILY SNAPSHOT"])
        sheet.append([slot.get("time", "") for slot in snapshot])
        sheet.append([slot.get("condition", "") for slot in snapshot])
        sheet.append([slot.get("temp_c", "") for slot in snapshot])
        sheet.append([])

    weather_obs = report.get("weather_observations") or []
    if weather_obs:
        evidence_by_id = {e.id: e for e in (report.get("evidence") or [])}
        sheet.append(["RAIN RECORDS"])
        sheet.append(["Start", "Finish", "Delay?", "Photo", "Comments"])
        for row in weather_obs:
            photo_evidence = evidence_by_id.get(_get(row, "evidence_id"))
            sheet.append(
                [
                    str(_get(row, "start_time") or ""),
                    str(_get(row, "end_time") or ""),
                    "Yes" if _get(row, "caused_delay") else "No",
                    (photo_evidence.caption or photo_evidence.filename) if photo_evidence else "",
                    _get(row, "comments"),
                ]
            )
        sheet.append([])

    notes_fields = [
        ("Work completed", report.get("work_completed")),
        ("Delays", report.get("delays")),
        ("Engineer's instruction", report.get("engineer_instruction")),
        ("Tomorrow's plan", report.get("tomorrow_plan")),
        ("Remarks", report.get("remarks")),
    ]
    if any(value for _, value in notes_fields):
        sheet.append(["NOTES"])
        for label, value in notes_fields:
            if value:
                sheet.append([label, value])
        sheet.append([])

    manpower = report.get("manpower_entries") or []
    if manpower:
        sheet.append(
            [f"MANPOWER LOG - {report.get('total_workers', 0)} Workers | {report.get('total_man_hours', 0)} Man Hours"]
        )
        sheet.append(["Company", "Trade", "Position", "Workers", "Hours", "Man Hours"])
        for row in manpower:
            hours = _get(row, "hours") or 0
            workers = _get(row, "workers_count") or 0
            sheet.append(
                [
                    _get(row, "company"),
                    _get(row, "trade"),
                    _get(row, "position"),
                    workers,
                    float(hours),
                    round(float(hours) * workers, 1),
                ]
            )
        sheet.append([])

    equipment = report.get("equipment_entries") or []
    if equipment:
        sheet.append(["EQUIPMENT LOG"])
        sheet.append(["Equipment", "Type", "Hrs Operating", "Hrs Idle", "Inspected?", "Location"])
        for row in equipment:
            sheet.append(
                [
                    _get(row, "equipment_name"),
                    _get(row, "equipment_type"),
                    _get(row, "hours_operating"),
                    _get(row, "hours_idle"),
                    "Yes" if _get(row, "inspected") else "No",
                    _get(row, "location"),
                ]
            )
        sheet.append([])

    delivery = report.get("delivery_entries") or []
    if delivery:
        sheet.append(["DELIVERY LOG"])
        sheet.append(["Time", "Delivered From", "Tracking No.", "Contents"])
        for row in delivery:
            sheet.append(
                [
                    str(_get(row, "delivery_time") or ""),
                    _get(row, "delivered_from"),
                    _get(row, "tracking_number"),
                    _get(row, "contents"),
                ]
            )
        sheet.append([])

    inspection = report.get("inspection_entries") or []
    if inspection:
        sheet.append(["INSPECTION LOG"])
        sheet.append(["Start", "End", "Type", "Entity", "Inspector", "Location"])
        for row in inspection:
            sheet.append(
                [
                    str(_get(row, "start_time") or ""),
                    str(_get(row, "end_time") or ""),
                    _get(row, "inspection_type"),
                    _get(row, "inspecting_entity"),
                    _get(row, "inspector_name"),
                    _get(row, "location_area"),
                ]
            )
        sheet.append([])

    hse = report.get("hse_entries") or []
    if hse:
        sheet.append(["HSE LOG"])
        sheet.append(["Time", "Category", "Description", "Action Taken", "Reported By"])
        for row in hse:
            sheet.append(
                [
                    str(_get(row, "entry_time") or ""),
                    _get(row, "category"),
                    _get(row, "description"),
                    _get(row, "action_taken"),
                    _get(row, "reported_by"),
                ]
            )
        sheet.append([])

    visitors = report.get("visitor_entries") or []
    if visitors:
        sheet.append(["VISITOR LOG"])
        sheet.append(["Time In", "Time Out", "Visitor", "Company", "Purpose", "Host"])
        for row in visitors:
            sheet.append(
                [
                    str(_get(row, "time_in") or ""),
                    str(_get(row, "time_out") or ""),
                    _get(row, "visitor_name"),
                    _get(row, "company"),
                    _get(row, "purpose"),
                    _get(row, "host_name"),
                ]
            )
        sheet.append([])

    evidence = report.get("evidence") or []
    if evidence:
        sheet.append([f"PHOTOS ({len(evidence)})"])
        sheet.append(["Filename", "Caption", "Category"])
        for item in evidence:
            sheet.append([item.filename, item.caption or "", item.category or ""])


def _sheet_title_for_date(diary_date, used_titles: set) -> str:
    """Excel sheet names are capped at 31 chars and must be unique within
    a workbook - diary_date alone (YYYY-MM-DD, 10 chars) fits comfortably,
    but de-dupe defensively in case of same-day duplicate rows."""
    base = str(diary_date)[:31]
    title = base
    suffix = 2
    while title in used_titles:
        title = f"{base[:28]}-{suffix}"
        suffix += 1
    used_titles.add(title)
    return title


def generate_daily_log_excel(report: dict):
    """Single-day Excel export, matching the reference site-log
    template's section layout - see _write_daily_log_sheet."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _sheet_title_for_date(report["diary_date"], set())
    _write_daily_log_sheet(sheet, report)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def generate_project_daily_log_compilation_excel(reports: list, project_name: str | None = None):
    """
    Every Daily Log for a project (optionally date-range filtered), one
    sheet per day - the Excel counterpart to
    pdf_service.generate_project_daily_log_compilation_pdf.
    """
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["Daily Log Compilation"])
    if project_name:
        summary_sheet.append(["Project", project_name])
    summary_sheet.append(["Days included", len(reports)])

    used_titles = {"Summary"}
    for report in reports:
        sheet = workbook.create_sheet(_sheet_title_for_date(report["diary_date"], used_titles))
        _write_daily_log_sheet(sheet, report)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output